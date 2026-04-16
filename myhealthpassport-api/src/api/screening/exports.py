import csv
import io
import json
import re
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import Query, Depends
from fastapi.responses import StreamingResponse

from src.core.manager import get_current_user
from src.models.student_models import Students, SmartScaleData
from src.models.screening_models import NutritionScreening, BehaviouralScreening, DentalScreening, EyeScreening
from src.models.other_models import ClinicalRecomendations, ClinicalFindings
from . import router


_ROMAN = {
    "I": "1", "II": "2", "III": "3", "IV": "4", "V": "5", "VI": "6",
    "VII": "7", "VIII": "8", "IX": "9", "X": "10", "XI": "11", "XII": "12",
}


def _norm_class(val: str) -> str:
    """Normalize class value to a plain digit string for comparison."""
    if not val:
        return ""
    v = re.sub(r'(?i)^\s*class\s*', '', str(val).strip()).strip()
    if v.upper() in _ROMAN:
        return _ROMAN[v.upper()]
    v = re.sub(r'(?i)(st|nd|rd|th)\b', '', v).strip()
    v = re.sub(r'(?i)\s*(class|grade|std|standard)\s*', '', v).strip()
    if v.isdigit():
        return v
    if v.upper() in _ROMAN:
        return _ROMAN[v.upper()]
    return str(val).strip()


def _make_student_name(student) -> str:
    parts = [student.first_name, student.middle_name, student.last_name]
    return " ".join(p for p in parts if p and p.strip()).strip()


def _extract_answers(questions_data, question_type: str) -> str:
    """Pull answer list for a given question_type out of a report's questions_data."""
    if not questions_data:
        return ""
    if isinstance(questions_data, str):
        try:
            questions_data = json.loads(questions_data)
        except Exception:
            return ""
    for qd in questions_data:
        if isinstance(qd, dict) and qd.get("question_type") == question_type:
            answers = qd.get("answers", [])
            if isinstance(answers, list):
                return "; ".join(str(a) for a in answers if a)
            return str(answers)
    return ""


async def _get_students(school_id: int, class_name: Optional[str], section: Optional[str]):
    """Return students filtered by school / class / section, ordered for export."""
    all_students = await Students.filter(
        school_students__school_id=school_id, is_deleted=False
    ).order_by("class_room", "section", "roll_no")

    if class_name:
        norm = _norm_class(class_name)
        all_students = [s for s in all_students if _norm_class(s.class_room) == norm]
    if section:
        sec_upper = section.strip().upper()
        all_students = [s for s in all_students if s.section.strip().upper() == sec_upper]

    return all_students


def _filename_suffix(class_name, section) -> str:
    label = f"Class{class_name}" if class_name else "AllClasses"
    if section:
        label += section.upper()
    return label


# ─────────────────────────────────────────────────────────────────────────────
# 1. Nutrition Screening Checklist
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/nutrition-checklist")
async def export_nutrition_checklist(
    school_id: int = Query(..., description="School ID"),
    class_name: Optional[str] = Query(None, description="Class (e.g. 9)"),
    section: Optional[str] = Query(None, description="Section (e.g. A)"),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    screenings = (
        await NutritionScreening.filter(student_id__in=student_ids, is_deleted=False)
        .order_by("-created_at")
        .all()
    )
    # keep only the most-recent entry per student
    seen = set()
    screening_map = {}
    for ns in screenings:
        if ns.student_id not in seen:
            screening_map[ns.student_id] = ns
            seen.add(ns.student_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Roll No", "Name", "Gender", "Class", "Section",
        "Eyes", "Hair", "Mouth/Lips", "Skin", "Nails", "Teeth",
        "General Signs", "Bone & Muscle", "Notes", "Saved At",
    ])

    for student in students:
        ns = screening_map.get(student.id)
        writer.writerow([
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
            ns.eyes if ns else "",
            ns.hair if ns else "",
            ns.mouth_lips if ns else "",
            ns.skin if ns else "",
            ns.nails if ns else "",
            ns.teeth if ns else "",
            ns.general_signs if ns else "",
            ns.bone_muscle if ns else "",
            ns.note if ns else "",
            ns.created_at.strftime("%Y-%m-%d %H:%M") if ns else "",
        ])

    filename = f"nutrition-checklist_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 2. Nutritional Analyst Report
# ─────────────────────────────────────────────────────────────────────────────
REPORT_TYPES = [
    "Physical Screening Report",
    "Questionnaire Reports",
    "Nutrition Deficiency Report",
    "Lab Reports",
]


@router.get("/export/nutrition-analysis")
async def export_nutrition_analysis(
    school_id: int = Query(...),
    class_name: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    recs = await ClinicalRecomendations.filter(
        student_id__in=student_ids, is_deleted=False
    ).all()

    # Group: student_id → {report_type → record}
    recs_by_student: dict = {}
    for r in recs:
        recs_by_student.setdefault(r.student_id, {})[r.report_type] = r

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["Roll No", "Name", "Gender", "Class", "Section"]
    for rt in REPORT_TYPES:
        short = rt.replace(" Report", "").replace(" Reports", "")
        headers += [
            f"{short} - Good Outcomes",
            f"{short} - Areas of Concern",
            f"{short} - Summary",
            f"{short} - Status",
        ]
    headers += ["Common Status", "Common Summary", "Clinical Notes", "Analysis Status", "Last Updated"]
    writer.writerow(headers)

    for student in students:
        student_recs = recs_by_student.get(student.id, {})
        any_rec = next(iter(student_recs.values()), None)

        row = [
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
        ]

        for rt in REPORT_TYPES:
            rec = student_recs.get(rt)
            qdata = rec.questions_data if rec else None
            row += [
                _extract_answers(qdata, "Good Outcomes"),
                _extract_answers(qdata, "Areas of Concern"),
                rec.summary or "" if rec else "",
                rec.status or "" if rec else "",
            ]

        row += [
            any_rec.common_status or "" if any_rec else "",
            any_rec.common_summary or "" if any_rec else "",
            any_rec.clinical_notes or "" if any_rec else "",
            "Complete" if (any_rec and any_rec.analysis_status) else "Incomplete",
            any_rec.updated_at.strftime("%Y-%m-%d %H:%M") if any_rec else "",
        ]
        writer.writerow(row)

    filename = f"nutrition-analysis_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3. Psychology Analysis
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/psychology-analysis")
async def export_psychology_analysis(
    school_id: int = Query(...),
    class_name: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    findings_qs = await ClinicalFindings.filter(
        student_id__in=student_ids, is_deleted=False
    ).all()
    # keep most recent per student
    seen = set()
    findings_map = {}
    for cf in findings_qs:
        if cf.student_id not in seen:
            findings_map[cf.student_id] = cf
            seen.add(cf.student_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Roll No", "Name", "Gender", "Class", "Section",
        "Good Strengths - Findings", "Good Strengths - Remarks",
        "Need Attention - Findings", "Need Attention - Remarks",
        "Status", "Summary", "Clinical Notes & Recommendations",
        "Analysis Status", "Last Updated",
    ])

    def _flatten_findings(data_list) -> tuple[str, str]:
        """Return (findings_str, remarks_str) from a list of {findings, remarks} dicts."""
        if not data_list:
            return ("", "")
        try:
            if isinstance(data_list, str):
                data_list = json.loads(data_list)
        except Exception:
            return ("", "")
        findings = "; ".join(item.get("findings", "") for item in data_list if item.get("findings"))
        remarks = "; ".join(item.get("remarks", "") for item in data_list if item.get("remarks"))
        return (findings, remarks)

    for student in students:
        cf = findings_map.get(student.id)
        good_findings, good_remarks = _flatten_findings(cf.findings_data if cf else None)
        need_findings, need_remarks = _flatten_findings(cf.need_attention_data if cf else None)

        writer.writerow([
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
            good_findings,
            good_remarks,
            need_findings,
            need_remarks,
            cf.status or "" if cf else "",
            cf.summary or "" if cf else "",
            cf.clinical_notes_recommendations or "" if cf else "",
            "Complete" if (cf and cf.analysis_status) else "Incomplete",
            cf.updated_at.strftime("%Y-%m-%d %H:%M") if cf else "",
        ])

    filename = f"psychology-analysis_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 4. Smart Scale Data
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/smart-scale")
async def export_smart_scale(
    school_id: int = Query(...),
    class_name: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    smart_data = (
        await SmartScaleData.filter(student_id__in=student_ids)
        .order_by("student_id", "-weighing_time")
        .all()
    )
    # latest entry per student
    seen = set()
    smart_map = {}
    for sd in smart_data:
        if sd.student_id not in seen:
            smart_map[sd.student_id] = sd
            seen.add(sd.student_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Roll No", "Name", "Gender", "Class", "Section",
        "Height (cm)", "Age (years)", "Weighing Time", "Body Weight (kg)", "BMI",
        "Body Fat Rate (%)", "Fat Content (kg)", "Lean Body Mass (kg)",
        "Muscle Mass (kg)", "Muscle Rate (%)", "Skeletal Muscle Mass (kg)",
        "Skeletal Muscle Mass Index", "Bone Mass (kg)", "Protein Content (kg)",
        "Internal Protein Rate (%)", "Visceral Fat Level", "Subcutaneous Fat Volume (kg)",
        "Subcutaneous Fat Rate (%)", "Water Content (kg)", "Body Moisture (%)",
        "Standard Body Weight (kg)", "Ideal Weight (kg)", "Physical Score",
        "Physical Age", "Body Type", "Heart Rate (bpm)", "Health Level",
        "Left Hand Fat Mass (kg)", "Right Hand Fat Mass (kg)",
        "Left Foot Fat Mass (kg)", "Right Foot Fat Mass (kg)", "Trunk Fat Mass (kg)",
        "Left Hand Muscle Mass (kg)", "Right Hand Muscle Mass (kg)",
        "Left Foot Muscle Mass (kg)", "Right Foot Muscle Mass (kg)", "Trunk Muscle Mass (kg)",
        "Saved At",
    ])

    for student in students:
        sd = smart_map.get(student.id)
        writer.writerow([
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
            sd.height_cm if sd else "",
            sd.age_years if sd else "",
            sd.weighing_time.strftime("%Y-%m-%d %H:%M") if (sd and sd.weighing_time) else "",
            sd.body_weight_kg if sd else "",
            sd.bmi if sd else "",
            sd.body_fat_rate_percent if sd else "",
            sd.fat_content_kg if sd else "",
            sd.lean_body_mass_kg if sd else "",
            sd.muscle_mass_kg if sd else "",
            sd.muscle_rate_percent if sd else "",
            sd.skeletal_muscle_mass_kg if sd else "",
            sd.skeletal_muscle_mass_index if sd else "",
            sd.bone_mass_kg if sd else "",
            sd.protein_content_kg if sd else "",
            sd.internal_protein_rate_percent if sd else "",
            sd.visceral_fat_level if sd else "",
            sd.subcutaneous_fat_volume_kg if sd else "",
            sd.subcutaneous_fat_rate if sd else "",
            sd.water_content_kg if sd else "",
            sd.body_moisture_content_percent if sd else "",
            sd.standard_body_weight if sd else "",
            sd.ideal_weight if sd else "",
            sd.physical_score if sd else "",
            sd.physical_age if sd else "",
            sd.body_type if sd else "",
            sd.heart_rate_beats_min if sd else "",
            sd.health_level if sd else "",
            sd.left_hand_fat_mass_kg if sd else "",
            sd.right_hand_fat_mass_kg if sd else "",
            sd.left_foot_fat_mass_kg if sd else "",
            sd.right_foot_fat_mass_kg if sd else "",
            sd.trunk_fat_mass_kg if sd else "",
            sd.left_hand_muscle_mass_kg if sd else "",
            sd.right_hand_muscle_mass_kg if sd else "",
            sd.left_foot_muscle_mass_kg if sd else "",
            sd.right_foot_muscle_mass_kg if sd else "",
            sd.trunk_muscle_mass_kg if sd else "",
            sd.weighing_time.strftime("%Y-%m-%d %H:%M") if (sd and sd.weighing_time) else "",
        ])

    filename = f"smart-scale_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 5. Psychology Screening Checklist
# ─────────────────────────────────────────────────────────────────────────────
_BS_CATEGORIES = [
    ("socialisation",      "Socialisation"),
    ("communication",      "Communication"),
    ("play_behaviour",     "Play Behaviour"),
    ("interaction",        "Interaction"),
    ("anxiety_withdrawal", "Anxiety/Withdrawal"),
    ("problem_behaviour",  "Problem Behaviour"),
]

_BS_RECOMMENDATIONS = [
    ("no_immediate_concern",    "No Immediate Concern"),
    ("continue_observation",    "Continue Observation"),
    ("discuss_with_teacher",    "Discuss with Teacher"),
    ("schedule_parent_meeting", "Parent Meeting"),
    ("refer_for_counselling",   "Refer for Counselling"),
    ("recommend_evaluation",    "Recommend Evaluation"),
]


def _bs_flag(mapping, key) -> str:
    """Return 'Yes'/'No'/'' from a JSON dict (may be str or dict)."""
    if not mapping:
        return ""
    if isinstance(mapping, str):
        try:
            mapping = json.loads(mapping)
        except Exception:
            return ""
    val = mapping.get(key)
    if val is True:
        return "Yes"
    if val is False:
        return "No"
    return str(val) if val is not None else ""


def _bs_items(category_data) -> str:
    """Flatten a category JSON dict into 'Item: Yes/No; ...' string."""
    if not category_data:
        return ""
    if isinstance(category_data, str):
        try:
            category_data = json.loads(category_data)
        except Exception:
            return ""
    parts = []
    for item, val in category_data.items():
        answer = "Yes" if val is True else ("No" if val is False else str(val))
        parts.append(f"{item}: {answer}")
    return "; ".join(parts)


@router.get("/export/psychology-checklist")
async def export_psychology_checklist(
    school_id: int = Query(...),
    class_name: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    screenings = (
        await BehaviouralScreening.filter(student_id__in=student_ids, is_deleted=False)
        .order_by("-created_at")
        .all()
    )
    # keep most-recent entry per student
    seen: set = set()
    screening_map: dict = {}
    for bs in screenings:
        if bs.student_id not in seen:
            screening_map[bs.student_id] = bs
            seen.add(bs.student_id)

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["Roll No", "Name", "Gender", "Class", "Section"]
    for _, cat_label in _BS_CATEGORIES:
        headers.append(f"{cat_label} - Concern")
        headers.append(f"{cat_label} - Checklist Items")
    for _, rec_label in _BS_RECOMMENDATIONS:
        headers.append(rec_label)
    headers += ["Gross Motor Skills", "Notes", "Next Followup", "Screening Status", "Saved At"]
    writer.writerow(headers)

    for student in students:
        bs = screening_map.get(student.id)
        row = [
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
        ]
        for cat_key, _ in _BS_CATEGORIES:
            cat_data = getattr(bs, cat_key, None) if bs else None
            row.append(_bs_flag(bs.summary_concerns if bs else None, cat_key))
            row.append(_bs_items(cat_data))
        for rec_key, _ in _BS_RECOMMENDATIONS:
            row.append(_bs_flag(bs.recommendations if bs else None, rec_key))
        row += [
            bs.gross_motor_skills if bs else "",
            bs.note if bs else "",
            bs.next_followup if bs else "",
            "Complete" if (bs and bs.screening_status) else "Incomplete",
            bs.created_at.strftime("%Y-%m-%d %H:%M") if bs else "",
        ]
        writer.writerow(row)

    filename = f"psychology-checklist_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 5. Dental Screening Export
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/dental-screening")
async def export_dental_screening(
    school_id: int = Query(..., description="School ID"),
    class_name: Optional[str] = Query(None, description="Class (e.g. 7)"),
    section: Optional[str] = Query(None, description="Section (e.g. A)"),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    screenings = (
        await DentalScreening.filter(student_id__in=student_ids, is_deleted=False)
        .order_by("-created_at")
        .all()
    )
    # Keep only the most-recent entry per student
    seen: set = set()
    screening_map: dict = {}
    for ds in screenings:
        if ds.student_id not in seen:
            screening_map[ds.student_id] = ds
            seen.add(ds.student_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Roll No", "Name", "Gender", "Class", "Section",
        "Patient Concern", "Oral Examination", "Examination Note",
        "Diagnosis", "Treatment Recommendations", "Treatment Recommendations Note",
        "Report Summary", "Next Follow-up", "Screening Status", "Saved At",
    ])

    for student in students:
        ds = screening_map.get(student.id)
        writer.writerow([
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
            ds.patient_concern if ds else "",
            ds.oral_examination if ds else "",
            ds.examination_note if ds else "",
            ds.diagnosis if ds else "",
            ds.treatment_recommendations if ds else "",
            ds.treatment_recommendations_note if ds else "",
            ds.report_summary if ds else "",
            ds.next_followup if ds else "",
            "Complete" if (ds and ds.screening_status) else "Incomplete",
            ds.created_at.strftime("%Y-%m-%d %H:%M") if ds else "",
        ])

    filename = f"dental-screening_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 6. Vision (Eye) Screening Export
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/vision-screening")
async def export_vision_screening(
    school_id: int = Query(..., description="School ID"),
    class_name: Optional[str] = Query(None, description="Class (e.g. 7)"),
    section: Optional[str] = Query(None, description="Section (e.g. A)"),
    current_user: dict = Depends(get_current_user),
):
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    screenings = (
        await EyeScreening.filter(student_id__in=student_ids, is_deleted=False)
        .order_by("-created_at")
        .all()
    )
    # Keep only the most-recent entry per student
    seen: set = set()
    screening_map: dict = {}
    for es in screenings:
        if es.student_id not in seen:
            screening_map[es.student_id] = es
            seen.add(es.student_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Roll No", "Name", "Gender", "Class", "Section",
        "Patient Concern", "Vision Left Eye", "Vision Right Eye",
        "Additional Findings", "Recommendations", "Report Summary",
        "Next Follow-up", "Screening Status", "Saved At",
    ])

    for student in students:
        es = screening_map.get(student.id)
        writer.writerow([
            student.roll_no,
            _make_student_name(student),
            student.gender,
            student.class_room,
            student.section,
            es.patient_concern if es else "",
            es.vision_lefteye_res if es else "",
            es.vision_righteye_res if es else "",
            es.additional_find if es else "",
            es.recommendations if es else "",
            es.report_summary if es else "",
            es.next_followup if es else "",
            "Complete" if (es and es.screening_status) else "Incomplete",
            es.created_at.strftime("%Y-%m-%d %H:%M") if es else "",
        ])

    filename = f"vision-screening_{_filename_suffix(class_name, section)}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 7. All Reports — Multi-sheet Excel (.xlsx)
# ─────────────────────────────────────────────────────────────────────────────
def _style_header_row(ws, header_row: list, fill_hex: str):
    """Write and style a header row in a worksheet."""
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for col_idx, value in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


@router.get("/export/all-reports-excel")
async def export_all_reports_excel(
    school_id: int = Query(..., description="School ID"),
    class_name: Optional[str] = Query(None, description="Class (e.g. 7)"),
    section: Optional[str] = Query(None, description="Section (e.g. A)"),
    current_user: dict = Depends(get_current_user),
):
    """Export all screening data (Dental, Vision, Nutrition, Psychology) as a multi-sheet Excel file."""
    students = await _get_students(school_id, class_name, section)
    student_ids = [s.id for s in students]

    # ── Fetch all screening data ──────────────────────────────────────────────
    dental_qs = await DentalScreening.filter(student_id__in=student_ids, is_deleted=False).order_by("-created_at").all()
    vision_qs = await EyeScreening.filter(student_id__in=student_ids, is_deleted=False).order_by("-created_at").all()
    nutrition_qs = await NutritionScreening.filter(student_id__in=student_ids, is_deleted=False).order_by("-created_at").all()
    behaviour_qs = await BehaviouralScreening.filter(student_id__in=student_ids, is_deleted=False).order_by("-created_at").all()

    def _latest_map(records):
        seen, result = set(), {}
        for r in records:
            if r.student_id not in seen:
                result[r.student_id] = r
                seen.add(r.student_id)
        return result

    dental_map = _latest_map(dental_qs)
    vision_map = _latest_map(vision_qs)
    nutrition_map = _latest_map(nutrition_qs)
    behaviour_map = _latest_map(behaviour_qs)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Sheet 1: Dental ───────────────────────────────────────────────────────
    ws_dental = wb.create_sheet("Dental Screening")
    dental_headers = [
        "Roll No", "Name", "Gender", "Class", "Section",
        "Patient Concern", "Oral Examination", "Examination Note",
        "Diagnosis", "Treatment Recommendations", "Treatment Recommendations Note",
        "Report Summary", "Next Follow-up", "Screening Status", "Saved At",
    ]
    _style_header_row(ws_dental, dental_headers, "34C789")
    for student in students:
        ds = dental_map.get(student.id)
        ws_dental.append([
            student.roll_no, _make_student_name(student), student.gender, student.class_room, student.section,
            ds.patient_concern if ds else "",
            ds.oral_examination if ds else "",
            ds.examination_note if ds else "",
            ds.diagnosis if ds else "",
            ds.treatment_recommendations if ds else "",
            ds.treatment_recommendations_note if ds else "",
            ds.report_summary if ds else "",
            ds.next_followup if ds else "",
            "Complete" if (ds and ds.screening_status) else "Incomplete",
            ds.created_at.strftime("%Y-%m-%d %H:%M") if ds else "",
        ])

    # ── Sheet 2: Vision ───────────────────────────────────────────────────────
    ws_vision = wb.create_sheet("Vision Screening")
    vision_headers = [
        "Roll No", "Name", "Gender", "Class", "Section",
        "Patient Concern", "Vision Left Eye", "Vision Right Eye",
        "Additional Findings", "Recommendations", "Report Summary",
        "Next Follow-up", "Screening Status", "Saved At",
    ]
    _style_header_row(ws_vision, vision_headers, "F59E0B")
    for student in students:
        es = vision_map.get(student.id)
        ws_vision.append([
            student.roll_no, _make_student_name(student), student.gender, student.class_room, student.section,
            es.patient_concern if es else "",
            es.vision_lefteye_res if es else "",
            es.vision_righteye_res if es else "",
            es.additional_find if es else "",
            es.recommendations if es else "",
            es.report_summary if es else "",
            es.next_followup if es else "",
            "Complete" if (es and es.screening_status) else "Incomplete",
            es.created_at.strftime("%Y-%m-%d %H:%M") if es else "",
        ])

    # ── Sheet 3: Nutrition ────────────────────────────────────────────────────
    ws_nutrition = wb.create_sheet("Nutrition Screening")
    nutrition_headers = [
        "Roll No", "Name", "Gender", "Class", "Section",
        "Eyes", "Hair", "Mouth/Lips", "Skin", "Nails", "Teeth",
        "General Signs", "Bone & Muscle", "Notes", "Screening Status", "Saved At",
    ]
    _style_header_row(ws_nutrition, nutrition_headers, "10B981")
    for student in students:
        ns = nutrition_map.get(student.id)
        ws_nutrition.append([
            student.roll_no, _make_student_name(student), student.gender, student.class_room, student.section,
            ns.eyes if ns else "",
            ns.hair if ns else "",
            ns.mouth_lips if ns else "",
            ns.skin if ns else "",
            ns.nails if ns else "",
            ns.teeth if ns else "",
            ns.general_signs if ns else "",
            ns.bone_muscle if ns else "",
            ns.note if ns else "",
            "Complete" if (ns and ns.screening_status) else "Incomplete",
            ns.created_at.strftime("%Y-%m-%d %H:%M") if ns else "",
        ])

    # ── Sheet 4: Psychology / Behavioural ─────────────────────────────────────
    ws_psych = wb.create_sheet("Psychology Screening")
    psych_headers = [
        "Roll No", "Name", "Gender", "Class", "Section",
        "Gross Motor Skills", "Notes", "Next Follow-up", "Screening Status", "Saved At",
    ]
    _style_header_row(ws_psych, psych_headers, "8B5CF6")
    for student in students:
        bs = behaviour_map.get(student.id)
        ws_psych.append([
            student.roll_no, _make_student_name(student), student.gender, student.class_room, student.section,
            bs.gross_motor_skills if bs else "",
            bs.note if bs else "",
            bs.next_followup if bs else "",
            "Complete" if (bs and bs.screening_status) else "Incomplete",
            bs.created_at.strftime("%Y-%m-%d %H:%M") if bs else "",
        ])

    # ── Auto-size columns on all sheets ──────────────────────────────────────
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Stream response ───────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"All-Reports_{_filename_suffix(class_name, section)}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
